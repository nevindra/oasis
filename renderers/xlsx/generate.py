#!/usr/bin/env python3
"""Generate XLSX from a JSON specification using openpyxl."""

import argparse
import json
import sys
from pathlib import Path

from openpyxl import Workbook
from openpyxl.chart import BarChart, LineChart, PieChart, AreaChart, ScatterChart, DoughnutChart, Reference
from openpyxl.formatting.rule import ColorScaleRule, DataBarRule, CellIsRule
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment, numbers
from openpyxl.utils import get_column_letter


CHART_TYPES = {
    "bar": BarChart,
    "line": LineChart,
    "pie": PieChart,
    "area": AreaChart,
    "scatter": ScatterChart,
    "doughnut": DoughnutChart,
}

HEADER_FILL = PatternFill(start_color="1B2A4A", end_color="1B2A4A", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
THIN_BORDER = Border(
    bottom=Side(style="thin", color="CCCCCC"),
)


def build_sheet(wb, sheet_spec):
    """Create and populate a worksheet from spec."""
    ws = wb.create_sheet(title=sheet_spec["name"])

    columns = sheet_spec.get("columns", [])
    rows = sheet_spec.get("rows", [])

    # Write headers.
    for col_idx, col_def in enumerate(columns, 1):
        cell = ws.cell(row=1, column=col_idx, value=col_def["header"])
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center")
        if col_def.get("width"):
            ws.column_dimensions[get_column_letter(col_idx)].width = col_def["width"]

    # Write data rows.
    for row_idx, row_data in enumerate(rows, 2):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = THIN_BORDER
            # Apply number format from column spec.
            if col_idx <= len(columns) and columns[col_idx - 1].get("format"):
                cell.number_format = columns[col_idx - 1]["format"]

    # Apply formulas.
    for formula_spec in sheet_spec.get("formulas", []):
        cell_ref = formula_spec["cell"]
        ws[cell_ref] = formula_spec["formula"]
        # Apply format from corresponding column.
        from openpyxl.utils import column_index_from_string
        col_letter = "".join(c for c in cell_ref if c.isalpha())
        col_num = column_index_from_string(col_letter)
        if col_num <= len(columns) and columns[col_num - 1].get("format"):
            ws[cell_ref].number_format = columns[col_num - 1]["format"]

    # Freeze panes.
    if sheet_spec.get("freeze_panes"):
        ws.freeze_panes = sheet_spec["freeze_panes"]

    # Charts.
    for chart_spec in sheet_spec.get("charts", []):
        chart_type = chart_spec.get("type", "bar")
        chart_cls = CHART_TYPES.get(chart_type, BarChart)
        chart = chart_cls()
        chart.title = chart_spec.get("title", "")

        # Parse data range (e.g., "A1:C13").
        data_range = chart_spec.get("data_range", "")
        if data_range and ":" in data_range:
            parts = data_range.split(":")
            start_col = column_index_from_string("".join(c for c in parts[0] if c.isalpha()))
            end_col = column_index_from_string("".join(c for c in parts[1] if c.isalpha()))
            start_row = int("".join(c for c in parts[0] if c.isdigit()) or "1")
            end_row = int("".join(c for c in parts[1] if c.isdigit()) or str(len(rows) + 1))

            cats = Reference(ws, min_col=start_col, min_row=start_row + 1, max_row=end_row)
            for c in range(start_col + 1, end_col + 1):
                values = Reference(ws, min_col=c, min_row=start_row, max_row=end_row)
                chart.add_data(values, titles_from_data=True)
            chart.set_categories(cats)

        size = chart_spec.get("size", {})
        chart.width = size.get("width", 15)
        chart.height = size.get("height", 10)
        ws.add_chart(chart, chart_spec.get("position", "G2"))

    # Conditional formatting.
    for cf in sheet_spec.get("conditional_formatting", []):
        cf_type = cf.get("type", "color_scale")
        cell_range = cf["range"]

        if cf_type == "color_scale":
            rule = ColorScaleRule(
                start_type="min", start_color=cf.get("min_color", "F87171").lstrip("#"),
                end_type="max", end_color=cf.get("max_color", "34D399").lstrip("#"),
            )
            ws.conditional_formatting.add(cell_range, rule)

        elif cf_type == "data_bar":
            rule = DataBarRule(
                start_type="min", end_type="max",
                color=cf.get("color", "2D5F8A").lstrip("#"),
            )
            ws.conditional_formatting.add(cell_range, rule)


def main():
    parser = argparse.ArgumentParser(description="Generate XLSX from JSON spec")
    parser.add_argument("input", help="JSON spec file")
    parser.add_argument("output", help="Output XLSX path")
    args = parser.parse_args()

    with open(args.input) as f:
        spec = json.load(f)

    wb = Workbook()
    # Remove the default empty sheet.
    wb.remove(wb.active)

    for sheet_spec in spec.get("sheets", []):
        build_sheet(wb, sheet_spec)

    wb.save(args.output)
    print(str(Path(args.output).resolve()))


if __name__ == "__main__":
    main()
